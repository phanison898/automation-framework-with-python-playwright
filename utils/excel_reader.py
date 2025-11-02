import openpyxl
from pathlib import Path

class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path)

    def get_sheet(self, sheet_name: str):
        """Return sheet object"""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {self.file_path}")
        return self.workbook[sheet_name]

    def get_all_data(self, sheet_name: str):
        """Return all data as list of dicts (first row as headers)"""
        sheet = self.get_sheet(sheet_name)
        rows = list(sheet.iter_rows(values_only=True))
        headers = [h.strip() if isinstance(h, str) else h for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def get(self, sheet_name: str, key: str):
        """
        Return row data as dict for given 'key'
        Example: excel.get("login", "key") -> {'username': 'x', 'password': 'y'}
        """
        sheet_data = self.get_all_data(sheet_name)
        for row in sheet_data:
            if str(row.get("key")).strip() == str(key).strip():
                # Exclude 'key' itself
                return {k: v for k, v in row.items() if k != "key"}
        raise ValueError(f"key '{key}' not found in sheet '{sheet_name}'")
